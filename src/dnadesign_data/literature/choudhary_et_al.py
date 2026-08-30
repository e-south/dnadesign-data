"""
--------------------------------------------------------------------------------
dnadesign-data
dnadesign-data/src/dnadesign_data/literature/choudhary_et_al.py

Hydrates Choudhary et al. BaeR ChIP-exo peaks into sequences.

Module Author(s): Eric J. South
--------------------------------------------------------------------------------
"""

from __future__ import annotations

import argparse
import csv
import datetime
import hashlib
import json
import sys
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

from dnadesign_data.core.layout import default_data_root, literature_path

REQUIRED_COLUMNS = {"peak_start", "peak_end", "strand", "peak_id"}
SOURCE_DOI = "10.1128/mSystems.00980-20"
COORDINATE_SYSTEM = "0-based, end-exclusive"

DATASET_DIR = default_data_root() / literature_path("Choudhary_et_al")
DEFAULT_XLSX = DATASET_DIR / "raw" / "mSystems.00980-20-sd002.xlsx"
DEFAULT_FASTA = DATASET_DIR / "raw" / "NC_000913.3.fasta"
DEFAULT_OUT_TSV = DATASET_DIR / "processed" / "BaeR_binding_sites.tsv"
DEFAULT_OUT_FASTA = DATASET_DIR / "processed" / "BaeR_binding_sites.fasta"
DEFAULT_PROVENANCE = DATASET_DIR / "processed" / "provenance.json"
DEFAULT_SHEET = "BaeR"
DEFAULT_TF = "BaeR"
DEFAULT_ACCESSION = "NC_000913.3"


def reverse_complement(sequence):
    complement_table = str.maketrans("ACGTN", "TGCAN")
    return sequence.upper().translate(complement_table)[::-1]


def normalize_column_name(name):
    if name is None:
        return ""
    cleaned = str(name).strip()
    if cleaned.upper() == "S/N":
        return "sn_ratio"
    cleaned = cleaned.replace(" ", "_").replace("/", "_")
    return cleaned.lower()


def apply_buffer(start, end, buffer, genome_length):
    if buffer < 0:
        raise ValueError("buffer must be >= 0")
    new_start = start - buffer
    new_end = end + buffer
    clipped = False
    if new_start < 0:
        new_start = 0
        clipped = True
    if new_end > genome_length:
        new_end = genome_length
        clipped = True
    return new_start, new_end, clipped


def slice_sequence(genome, start, end, strand):
    if strand not in {"+", "-"}:
        raise ValueError(f"invalid strand: {strand}")
    sequence = genome[start:end]
    if strand == "-":
        return reverse_complement(sequence)
    return sequence


def file_sha256(path):
    sha256 = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(8192), b""):
            sha256.update(chunk)
    return sha256.hexdigest()


def download_genome(accession, destination):
    url = (
        "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
        f"?db=nuccore&id={accession}&rettype=fasta&retmode=text"
    )
    destination.parent.mkdir(parents=True, exist_ok=True)
    with urllib.request.urlopen(url) as response:
        content = response.read()
    destination.write_bytes(content)


def load_genome_sequence(fasta_path, accession):
    try:
        from pyfaidx import Fasta
    except ImportError as exc:
        raise ImportError(
            "pyfaidx is required. Install dnadesign-data with its runtime dependencies."
        ) from exc
    if not fasta_path.exists():
        raise FileNotFoundError(f"genome FASTA not found: {fasta_path}")
    fasta = Fasta(str(fasta_path), as_raw=True, sequence_always_upper=True)
    if accession not in fasta:
        fasta.close()
        raise KeyError(f"accession not found in FASTA: {accession}")
    sequence = str(fasta[accession])
    fasta.close()
    return sequence


def load_sheet_rows(xlsx_path, sheet_name):
    if not xlsx_path.exists():
        raise FileNotFoundError(f"xlsx not found: {xlsx_path}")
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"sheet not found: {sheet_name}")
    sheet = workbook[sheet_name]
    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter)
    headers = [normalize_column_name(name) for name in header_row]
    rows = []
    for row in rows_iter:
        if all(cell is None for cell in row):
            continue
        row_dict = {headers[idx]: value for idx, value in enumerate(row)}
        rows.append(row_dict)
    workbook.close()
    return rows


def dedupe_rows(rows, mode):
    if mode == "none":
        return rows
    seen = set()
    deduped = []
    for row in rows:
        if mode == "peak_id":
            key = row.get("peak_id")
        elif mode == "coords":
            key = (row.get("chrom"), row.get("start"), row.get("end"))
        else:
            raise ValueError(f"invalid dedupe mode: {mode}")
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def build_output_rows(rows, genome, tf, chrom, buffer):
    genome_length = len(genome)
    output_rows = []
    for row in rows:
        start = int(row.get("peak_start"))
        end = int(row.get("peak_end"))
        if start >= end:
            raise ValueError(f"start >= end for peak_id {row.get('peak_id')}")
        adjusted_start, adjusted_end, clipped = apply_buffer(
            start, end, buffer, genome_length
        )
        if adjusted_start >= adjusted_end:
            raise ValueError(f"start >= end after buffer for {row.get('peak_id')}")
        if adjusted_start < 0 or adjusted_end > genome_length:
            raise ValueError(f"coordinates out of bounds for {row.get('peak_id')}")
        strand = row.get("strand")
        sequence = slice_sequence(genome, adjusted_start, adjusted_end, strand)
        output_rows.append(
            {
                "tf": tf,
                "chrom": chrom,
                "start": adjusted_start,
                "end": adjusted_end,
                "strand": strand,
                "sequence": sequence,
                "peak_id": row.get("peak_id"),
                "sn_ratio": row.get("sn_ratio"),
                "operon": row.get("operon"),
                "locus_tag": row.get("locus_tag"),
                "gene_name": row.get("gene_name"),
                "distance": row.get("distance"),
                "type": row.get("type"),
                "length": adjusted_end - adjusted_start,
                "midpoint": (adjusted_start + adjusted_end) // 2,
                "source_doi": SOURCE_DOI,
                "clipped": clipped,
            }
        )
    return output_rows


def write_tsv(path, rows, fieldnames):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=fieldnames, delimiter="\t", lineterminator="\n"
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_fasta(path, rows, tf, chrom):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w") as handle:
        for row in rows:
            peak_id = row.get("peak_id")
            strand = row.get("strand")
            operon = row.get("operon")
            sn_ratio = row.get("sn_ratio")
            site_type = row.get("type")
            header = (
                f">{tf}|{peak_id}|{chrom}:{row.get('start')}-{row.get('end')}"
                f"|strand={strand}|operon={operon}|sn={sn_ratio}|type={site_type}"
            )
            handle.write(header + "\n")
            handle.write(row.get("sequence", "") + "\n")


def write_provenance(path, args, xlsx_path, fasta_path, total_rows, output_rows):
    record = {
        "timestamp_utc": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "input_xlsx": provenance_path(xlsx_path),
        "input_xlsx_sha256": file_sha256(xlsx_path),
        "genome_fasta": provenance_path(fasta_path),
        "genome_fasta_sha256": file_sha256(fasta_path),
        "genome_accession": args.genome_accession,
        "coordinate_system": COORDINATE_SYSTEM,
        "dedupe_mode": args.dedupe,
        "rows_read": total_rows,
        "rows_written": len(output_rows),
        "cli_args": provenance_args(args),
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(record, indent=2) + "\n")


def provenance_args(args):
    values = vars(args).copy()
    for key in ("xlsx", "genome_fasta", "out_tsv", "out_fasta", "out_provenance"):
        if key in values:
            values[key] = provenance_path(Path(values[key]))
    return values


def provenance_path(path):
    resolved = Path(path).resolve()
    try:
        return resolved.relative_to(default_data_root().resolve()).as_posix()
    except ValueError:
        return str(path)


def parse_args(argv):
    parser = argparse.ArgumentParser(
        description="Hydrate ChIP-exo peaks into strand-aware sequences."
    )
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--tf", default=DEFAULT_TF)
    parser.add_argument("--genome-fasta", default=str(DEFAULT_FASTA))
    parser.add_argument("--genome-accession", default=DEFAULT_ACCESSION)
    parser.add_argument("--chrom", default=DEFAULT_ACCESSION)
    parser.add_argument("--out-tsv", default=str(DEFAULT_OUT_TSV))
    parser.add_argument("--out-fasta", default=str(DEFAULT_OUT_FASTA))
    parser.add_argument("--out-provenance", default=str(DEFAULT_PROVENANCE))
    parser.add_argument("--buffer", type=int, default=0)
    parser.add_argument(
        "--dedupe", choices=["none", "peak_id", "coords"], default="none"
    )
    parser.add_argument("--download-genome", action="store_true")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    xlsx_path = Path(args.xlsx)
    fasta_path = Path(args.genome_fasta)
    if args.download_genome:
        download_genome(args.genome_accession, fasta_path)
    rows = load_sheet_rows(xlsx_path, args.sheet)
    normalized_columns = (
        {normalize_column_name(col) for col in rows[0]} if rows else set()
    )
    missing = REQUIRED_COLUMNS - normalized_columns
    if missing:
        raise KeyError(f"missing required columns: {sorted(missing)}")
    genome = load_genome_sequence(fasta_path, args.genome_accession)
    output_rows = build_output_rows(rows, genome, args.tf, args.chrom, args.buffer)
    output_rows = dedupe_rows(output_rows, args.dedupe)
    fieldnames = [
        "tf",
        "chrom",
        "start",
        "end",
        "strand",
        "sequence",
        "peak_id",
        "sn_ratio",
        "operon",
        "locus_tag",
        "gene_name",
        "distance",
        "type",
        "length",
        "midpoint",
        "source_doi",
        "clipped",
    ]
    write_tsv(Path(args.out_tsv), output_rows, fieldnames)
    write_fasta(Path(args.out_fasta), output_rows, args.tf, args.chrom)
    write_provenance(
        Path(args.out_provenance),
        args,
        xlsx_path,
        fasta_path,
        total_rows=len(rows),
        output_rows=output_rows,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
